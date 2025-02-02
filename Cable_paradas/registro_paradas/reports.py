from django.views.generic import TemplateView
from django.http.response import HttpResponse
from .models import StopRegistration, OperationTime, EventStopCode, TechnicalData
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ReporterExcelStop(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or StopRegistration.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        stop_reg = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Paradas Registradas"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="13678A", end_color="13678A", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["A1"] = "REPORTE PARADAS REGISTRADAS"
        ws.merge_cells("A1:M1")
        ws["A1"].font = title_font
        ws["A1"].alignment = center_alignment
        
        headers = ["FECHA", "MOTIVO DE DETENCION", "ESTACION", "HORA DE DETENCION", "HORA INICIO OPERACION",
                "DURACION (seg)", "CABINA 1", "CABINA 2", "EVENTO", "TURNO", "OBSERVACIONES", "INPUTABLE", "OPERADOR"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        # Modificacion para el event stop code

        for stop in stop_reg:
            # Proceso para resolver el stop_code_name
            stop_code_name = "Code no found"
            if stop.stop_code:
                try:
                    stop_code_obj = EventStopCode.objects.get(
                        event_type=stop.event_type,
                        object_id=stop.stop_code
                    )
                    content_type = stop_code_obj.content_type
                    related_object = content_type.get_object_for_this_type(pk=stop_code_obj.object_id)

                    if hasattr(related_object, 'description'):
                        stop_code_name = related_object.code + " - " + related_object.description
                    elif hasattr(related_object, 'event_name'):
                        stop_code_name = related_object.event_name
                    elif hasattr(related_object, 'detention_name'):
                        stop_code_name = related_object.detention_name
                    elif hasattr(related_object, 'speed_name'):
                        stop_code_name = related_object.speed_name
                    else:
                        stop_code_name = "Nombre no encontrado"
                except EventStopCode.DoesNotExist:
                    stop_code_name = "Code no found"

            # Agregar datos al Excel
            ws.cell(row=cont, column=1).value = str(stop.registration_date)
            ws.cell(row=cont, column=2).value = stop_code_name  # Usa el nombre resuelto
            ws.cell(row=cont, column=3).value = str(stop.station)
            ws.cell(row=cont, column=4).value = str(stop.start_date)
            ws.cell(row=cont, column=5).value = str(stop.end_date)
            ws.cell(row=cont, column=6).value = str(stop.stop_time)
            ws.cell(row=cont, column=7).value = str(stop.cabin)
            ws.cell(row=cont, column=8).value = str(stop.cabin2)
            ws.cell(row=cont, column=9).value = str(stop.event_type)
            ws.cell(row=cont, column=10).value = str(stop.shift)
            ws.cell(row=cont, column=11).value = str(stop.observation)
            ws.cell(row=cont, column=12).value = str(stop.inputable)
            ws.cell(row=cont, column=13).value = str(stop.operator)

            cont += 1

        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_paradas_registradas.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response
    


class ReporterExcelOperatorDay(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or OperationTime.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        ope_reg = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Jornada de operacion"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="13678A", end_color="13678A", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["A1"] = "REPORTE JORNADA OPERACIONES"
        ws.merge_cells("A1:G1")
        ws["A1"].font = title_font
        ws["A1"].alignment = center_alignment
        
        headers = ["FECHA", "HORA INICIO", "HOROMETRO INICIO", "HORA FIN", "HOROMETRO FIN", "TIEMPO TOTAL", "OPERADOR"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        # Itera sobre el queryset filtrado
        for ope in ope_reg:
            ws.cell(row= cont, column= 1).value = str(ope.date)
            ws.cell(row= cont, column= 2).value = str(ope.start_time)
            ws.cell(row= cont, column= 3).value = str(ope.horometer_start)
            ws.cell(row= cont, column= 4).value = str(ope.end_time)
            ws.cell(row= cont, column= 5).value = str(ope.horometer_end)
            ws.cell(row= cont, column= 6).value = str(ope.total_time)
            ws.cell(row= cont, column= 7).value = str(ope.operator)
            
            cont +=1
        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_jornada_operacion.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response
    


class ReporterExcelTechnicalData(TemplateView):
    def __init__(self, queryset = None):
        # Utiliza el queryset pasado en el constructor o todos los registros si no se pasa
        self.queryset = queryset or TechnicalData.objects.all()
    
    def get(self, request, *args, **kwargs):
        # Aqui utilizamos el queryset filtrado en lugar de obtener todos los registros
        tec_reg = self.queryset        
        
        # Crea el archivo excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos tecnicos"
        
        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="13678A", end_color="13678A", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        ws["A1"] = "REPORTE DATOS TECNICOS"
        ws.merge_cells("A1:P1")
        ws["A1"].font = title_font
        ws["A1"].alignment = center_alignment
        
        headers = ["FECHA", "HORA", "TEMP MOTOR VARIADO", "TEMP CUARTO VARIADOR", "TEMP REDUCTOR", "TEMP REDUCTOR PANEL (C°)", "PAR (%)",
                    "PRESION FS BAR", "PRESION FE BAR", "CARRO TENSOR TUNAL", "CARRO TENSOR PARAISO", "LONGITUD DE LINEA", "VIENTO P06", 
                    "VIENTO P16", "VIENTO P23", "OPERADOR"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        cont = 4

        # Itera sobre el queryset filtrado
        for tec in tec_reg:
            ws.cell(row= cont, column= 1).value = str(tec.date)
            ws.cell(row= cont, column= 2).value = str(tec.time)
            ws.cell(row= cont, column= 3).value = str(tec.temp_motor_variador)
            ws.cell(row= cont, column= 4).value = str(tec.temp_cuarto_variador)
            ws.cell(row= cont, column= 5).value = str(tec.temp_reductor)
            ws.cell(row= cont, column= 6).value = str(tec.temp_reductor_panel)
            ws.cell(row= cont, column= 7).value = str(tec.par)
            ws.cell(row= cont, column= 8).value = str(tec.presion_fs_bar)
            ws.cell(row= cont, column= 9).value = str(tec.presion_fe_bar)
            ws.cell(row= cont, column= 10).value = str(tec.carro_tensor_tunal)
            ws.cell(row= cont, column= 11).value = str(tec.carro_tensor_paraiso)
            ws.cell(row= cont, column= 12).value = str(tec.longitud_de_linea)
            ws.cell(row= cont, column= 13).value = str(tec.viento_p06)
            ws.cell(row= cont, column= 14).value = str(tec.viento_p16)
            ws.cell(row= cont, column= 15).value = str(tec.viento_p23)
            ws.cell(row= cont, column= 16).value = str(tec.operator)
            
            cont +=1
        
        # Ajustar el ancho de las columnas
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = ws.cell(row=3, column=col_num).column_letter
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
                    

        file_name = "Reporte_datos_tecnicos.xlsx"
        response = HttpResponse(content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        content = "attachment; filename = {0}".format(file_name)
        response["Content-Disposition"] = content
        wb.save(response)
        return response